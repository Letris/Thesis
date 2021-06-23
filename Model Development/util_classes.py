
import copy
import statistics
from dataclasses import InitVar, dataclass, field
from pickle import dump
from typing import Any, Dict, Set
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import scipy
import scipy.stats
import seaborn as sns
import shap
from imblearn.combine import SMOTETomek
from imblearn.over_sampling import SMOTE, RandomOverSampler
from imblearn.under_sampling import NearMiss, RandomUnderSampler
from mealy.error_analyzer import ErrorAnalyzer
from mealy.error_visualizer import ErrorVisualizer
from sklearn import neighbors, svm
from sklearn.calibration import CalibratedClassifierCV
from sklearn.decomposition import PCA
from sklearn.ensemble import (AdaBoostClassifier, GradientBoostingClassifier,
                              RandomForestClassifier, StackingClassifier)
from sklearn.feature_selection import RFECV, SelectKBest, mutual_info_classif
from sklearn.inspection import permutation_importance
from sklearn.linear_model import LogisticRegression, RidgeClassifier
from sklearn.metrics import (accuracy_score, balanced_accuracy_score,
                             cohen_kappa_score, confusion_matrix, f1_score,
                             precision_score, recall_score, roc_auc_score)
from sklearn.model_selection import train_test_split
from sklearn.multiclass import OneVsOneClassifier, OneVsRestClassifier
from sklearn.neural_network import MLPClassifier
from sklearn.pipeline import make_pipeline
from sklearn.preprocessing import (MinMaxScaler, PolynomialFeatures,
                                   StandardScaler, power_transform)
from sklearn.tree import DecisionTreeClassifier
from sklearn.metrics import plot_confusion_matrix


@dataclass
class Dataset:
    """Data class to hold the dataset
    """

    # the features
    X: Any

    # the labels
    y: Any

    # the label encoder/decoder
    label_encoder: Any

    # the parameter
    params: Any

    # the choice of samplers
    available_samplers: InitVar

    def __post_init__(self, available_samplers):
        """Define all parameters of the instance

        Args:
            available_samplers (class instance): the choice of samplers
        """
        # split the data into a train and test set
        self.X, self.X_TEST, self.y, self.y_TEST = train_test_split(
            self.X, self.y, test_size=0.1, stratify=self.y, random_state=33)

        # make copies of the data
        self.X_copy = copy.deepcopy(self.X)
        self.X_TEST_copy = copy.deepcopy(self.X_TEST)
        self.y_copy = copy.deepcopy(self.y)

        # find the number of samples in the data
        self.n_samples = len(self.X)

        # decode the labels for confusion matrices
        self.labels = self.decode_labels()

        # store the names of the columns
        self.columns = self.X.columns.to_list()

        # set the sampling method to use, if any
        if self.params.SAMPLING_METHOD is not None:
            strategy_dict = {key: value for key, value in available_samplers.__dict__.items(
            ) if not key.startswith('__') and not callable(key)}
            if self.params.SAMPLING_METHOD == 'both':
                self.sampler = strategy_dict['over']
                self.sampler = strategy_dict['under']
                self.sampler = 'both'
            else:
                self.sampler = strategy_dict[self.params.SAMPLING_METHOD]
        else:
            self.sampler = None

    def decode_labels(self):
        """Decode the labels from int to original description.

        Returns:
            array: decoded labels
        """
        return self.label_encoder.inverse_transform(self.y)

    def preprocess(self):
        """Preprocess the data if steps were defiend
        """
        # standardize the data
        if self.params.STANDARDIZE_DATA:
            self.standardize_all_data()

        # normalize the data
        if self.params.NORMALIZE_DATA:
            self.normalize_all_data()

        # reduce the dimensions of the data with PCA
        if self.params.PCA_DATA:
            self.pca_all_data()

        # increase the dimensionality of the data with polynomial features
        if self.params.POLY_DATA:
            self.poly_all_data()

        # resample the data
        if self.sampler is not None:
            self.resample()

    def pca_all_data(self):
        """Apply PCA to the entire dataset

        Returns:
            numpy array: the dataset after PCA
        """
        self.pca_transformer = PCA(self.params.PCA_VARIANCE).fit(self.X)
        dump(self.pca_transformer, open(
            f'Results/Transformers/pca_{self.params.DATA_TYPE}_{self.params.PREDICTION_TYPE}.pkl', 'wb'))
        pca_data = self.pca_transformer.transform(self.X)
        self.X = pca_data
        self.X_TEST = self.pca_transformer.transform(self.X_TEST)
        return pca_data

    def standardize_all_data(self):
        """Standardize the dataset

        Returns:
            numpy array: the dataset after standardization
        """
        self.standardizer = StandardScaler().fit(self.X)
        dump(self.standardizer, open(
            f'Results/Transformers/standardizer_{self.params.DATA_TYPE}_{self.params.PREDICTION_TYPE}.pkl', 'wb'))
        standardized_data = self.standardizer.transform(self.X)
        self.X = standardized_data
        self.X_TEST = self.standardizer.transform(self.X_TEST)
        return standardized_data

    def normalize_all_data(self):
        """Normalize the entire dataset

        Returns:
            numpy array: the dataset after normalization
        """
        self.normalizer = MinMaxScaler().fit(self.X)
        dump(self.normalizer, open(
            f'Results//Transformers/normalizer_{self.params.DATA_TYPE}_{self.params.PREDICTION_TYPE}.pkl', 'wb'))
        normalized_data = self.normalizer.transform(self.X)
        self.X = normalized_data
        self.X_TEST = self.normalizer.transform(self.X_TEST)
        return normalized_data

    def poly_all_data(self):
        """Generate polynomial features

        Returns:
            numpy array: the dataset with added polynomial features
        """
        self.polynomial_transformer = PolynomialFeatures(
            self.params.AMT_POLY, interaction_only=True).fit(self.X)
        dump(self.polynomial_transformer, open(
            f'Results/Transformers/polynomial_{self.params.DATA_TYPE}_{self.params.PREDICTION_TYPE}.pkl', 'wb'))
        poly_data = self.polynomial_transformer.transform(self.X)
        self.X = poly_data
        self.X_TEST = self.polynomial_transformer.transform(self.X_TEST)
        return poly_data

    def resample(self):
        """Resample the data with according to the chosen sampling technique
        """
        # if oversampling and undersampling are combined
        if isinstance(self.sampler, str):
            # oversample the minority class
            self.X, self.y = self.oversampler.fit_resample(self.X, self.y)
            # undersample the majority class
            self.X, self.y = self.undersampler.fit_resample(self.X, self.y)
        else:
            # apply sampling method
            self.X, self.y = self.sampler.fit_resample(self.X, self.y)

    def save(self):
        """Store the data in excel
        """
        features = self.polynomial_transformer.get_feature_names(
            self.X_copy.columns)
        train_data = pd.DataFrame(self.X, columns=features)
        train_data.drop(['1'], axis=1, inplace=True)
        train_data.to_excel('Results/Transformers/train_data.xlsx')


@dataclass
class ConfidenceInterval:
    """Data class to hold confidence intervals
    """
    # the metric for which the confidence interval is stored
    metric: str
    # the mean of the metric scores
    mean: float
    # left limit of the confidence interval
    left_limit: float
    # right limit of the confidence interval
    right_limit: float


@dataclass
class Results:
    """Class to hold and process the results of a model
    """
    # the name of the model for which the results are stored
    model_name: str

    # lists that will hold the metric scores
    accuracy: list = field(default_factory=list)
    balanced_accuracy: list = field(default_factory=list)
    f1: list = field(default_factory=list)
    precision: list = field(default_factory=list)
    recall: list = field(default_factory=list)
    auc: list = field(default_factory=list)
    kappa: list = field(default_factory=list)

    # list that will hold the confusion matrices
    confusion_matrices: list = field(default_factory=list)

    # set with the available metric scores
    available_metrics: Set[str] = field(default_factory=lambda: {
                                        'accuracy', 'balanced_accuracy', 'f1', 'precision', 'recall', 'auc', 'kappa'})

    def add_scores(self, params, y_test, y_pred, y_proba):
        """Compute metric scores and add them to the lists holding the scores

        Args:
            params (class instance): parameters
            y_test (numpy array): true labels
            y_pred (numpy array): predicted labels
            y_proba (numpy array): predicted class probabilities
        """

        # compute the metrics scores and add them to the lists holding the scores
        self.accuracy.append(accuracy_score(y_test, y_pred))
        self.balanced_accuracy.append(balanced_accuracy_score(y_test, y_pred))
        self.f1.append(f1_score(y_test, y_pred, average='weighted'))
        self.precision.append(precision_score(
            y_test, y_pred, average='weighted'))
        self.recall.append(recall_score(y_test, y_pred, average='weighted'))
        self.kappa.append(cohen_kappa_score(y_test, y_pred))

        # need predicted class probabilities for computation of multiclass auc
        if params.MULTICLASS:
            try:
                self.auc.append(roc_auc_score(y_test, y_proba,
                                average='weighted', multi_class='ovr'))
            except ValueError:
                self.auc.append(0)
        else:
            self.auc.append(roc_auc_score(y_test, y_pred, average='weighted'))

    def average_scores(self):
        """Average the scores held by the metric score lists
        """
        self.mean_accuracy = round(statistics.mean(self.accuracy), 2)
        self.mean_balanced_accuracy = round(
            statistics.mean(self.balanced_accuracy), 2)
        self.mean_f1 = round(statistics.mean(self.f1), 2)
        self.mean_auc = round(statistics.mean(self.auc), 2)
        self.mean_precision = round(statistics.mean(self.precision), 2)
        self.mean_recall = round(statistics.mean(self.recall), 2)
        self.mean_kappa = round(statistics.mean(self.kappa), 2)

    def scores_to_dict(self):
        """Transform the mean metric scores to a dictionary

        Returns:
            dict: {metric : mean score}
        """
        self.score_dict = {f'{self.model_name}': {
            'Accuracy': self.mean_accuracy,
            'Balanced accuracy': self.mean_balanced_accuracy,
            'Precision': self.mean_precision,
            'Recall': self.mean_recall,
            'F1': self.mean_f1,
            'AUC': self.mean_auc,
            'Kappa': self.mean_kappa
        }}

        return self.score_dict

    def show_scores(self):
        """Print the averaged scores
        """
        print(f'Results: {self.model_name}')
        print(f'Accuracy: {self.mean_accuracy}')
        print(f'Balanced Accuracy: {self.mean_balanced_accuracy}')
        print(f'F1 score: {self.mean_f1}')
        print(f'AUC: {self.mean_auc}')
        print(f'Precision: {self.mean_precision}')
        print(f'Recall: {self.mean_recall}')
        print(f'Kappa: {self.mean_kappa}')

    def compute_confidence_intervals(self):
        """Compute the confidence interval for each metric based on the score lists
        """
        confidence_intervals = []

        # for each metric
        for metric in self.available_metrics:
            # obtain the scores
            scores = eval(f'self.{metric}')
            # compute the confidence interval
            m, left_limit, right_limit = self.compute_confidence_interval(
                scores, confidence=0.95)
            # store the confidence interval with a class instance
            confidence_interval = ConfidenceInterval(
                metric, m, left_limit, right_limit)
            # add the confidence interval to the list of confidence intervals
            confidence_intervals.append(confidence_interval)

        # store all confidence intervals
        self.confidence_intervals = confidence_intervals

    @staticmethod
    def compute_confidence_interval(data, confidence=0.95):
        """Compute the confidence interval of the given data.

        Args:
            data (list): list with the data for which to compute the confidence interval
        Returns:
            float, float, float: the mean, the left error, and the right error
        """
        a = 1.0 * np.array(data)
        n = len(a)
        m, se = np.mean(a), scipy.stats.sem(a)
        error_range = se * scipy.stats.t.ppf((1 + confidence) / 2, n-1)
        return m, m - error_range, m + error_range

    def add_confusion_matrix(self, y_test, y_pred):
        """Computes a confusion matrix and adds it to the list of matrices

        Args:
            y_test (numpy array): true labels
            y_pred (numpy array): predicted labels
        """
        CM = confusion_matrix(y_test, y_pred)
        self.confusion_matrices.append(CM)

    def plot_confusion_matrix_test(self, name, classifier, X_test, y_test, labels, data_type, ptype):
        """Plot a single confusion matrix

        Args:
            name (str): name of the classifier used to predict
            classifier (class instance): the classifier to predict X_test
            X_test (numpy array): the data to predict
            y_test (numpy array): the true labels of the data to predict
            labels (list): list of decoded labels for plotting
            data_type (str): knee or hip
            ptype (str): diagnosis or treatment
        """
        disp = plot_confusion_matrix(classifier, X_test, y_test, normalize='true',
                                     display_labels=labels,
                                     cmap=plt.cm.Blues)

        plt.savefig(f'cm_{data_type}_{ptype}_{name}.png')
        plt.close()

    def plot_confusion_matrix_validation(self, labels, data_type, ptype, figsize=(10, 10)):
        """Manually plot a confusion matrix

        Args:
            labels (list): list of decoded labels for plotting
            data_type (str): knee or hip
            ptype (str): diagnosis or hip
            figsize (tuple, optional): size of the figure. Defaults to (10,10).
        """

        if len(self.confusion_matrices) > 1:
            averaged_cm = np.average(self.confusion_matrices, axis=0)
        else:
            averaged_cm = self.confusion_matrices[0]

        cm_sum = np.sum(averaged_cm, axis=1, keepdims=True)
        cm_perc = averaged_cm / cm_sum.astype(float) * 100
        annot = np.empty_like(averaged_cm).astype(str)
        nrows, ncols = averaged_cm.shape
        for i in range(nrows):
            for j in range(ncols):
                c = averaged_cm[i, j]
                p = cm_perc[i, j]
                if i == j:
                    s = cm_sum[i]
                    annot[i, j] = '%.1f%%\n%d/%d' % (p, c, s)
                elif c == 0:
                    annot[i, j] = ''
                else:
                    annot[i, j] = '%.1f%%\n%d' % (p, c)

        cm = pd.DataFrame(averaged_cm, index=np.unique(
            labels), columns=np.unique(labels))
        cm.index.name = 'Actual'
        cm.columns.name = 'Predicted'
        fig, ax = plt.subplots(figsize=figsize)
        sns.heatmap(cm, cmap="YlGnBu", annot=annot, fmt='', ax=ax)
        plt.savefig(
            f'Results/Plots/confusion_matrices/{self.model_name}_{data_type}_{ptype}_cm.png')
        plt.close()

    def confidence_intervals_to_table(self, model_name, data_type, prediction_type, to_excel=True, to_latex=True):
        """retrieve the confidence intervals and store the table to excel and/or latex

        Args:
            model_name (str): the name of the model for which the intervals were computed
            data_type (str): knee or hip
            prediction_type (str): diagnosis or treatment
            to_excel (bool, optional): whether to store the intervals to excel. Defaults to True.
            to_latex (bool, optional): whether to store the intervals to latex. Defaults to True.
        """
        # dict to hold the intervals
        result_dict = {}
        result_dict[self.model_name] = {}

        # for each confidence interval
        for confidence_interval in self.confidence_intervals:
            # retrieve the name of the metric
            metric = confidence_interval.metric
            # retrieve and round the metric mean
            mean = round(confidence_interval.mean, 3)
            # retrieve and round the left and right limit
            left_limit = round(confidence_interval.left_limit, 3)
            right_limit = round(confidence_interval.right_limit, 3)
            # add the interval to the dict in string form
            result_dict[self.model_name][metric] = f'{mean} ({left_limit} - {right_limit})'

        # transform the dict with intervals to a table
        self.results_df_with_ci = pd.DataFrame.from_dict(
            result_dict, orient='index')
        # print the interval table
        print(self.results_df_with_ci)

        # possibly store the interval table to excel
        if to_excel:
            append_df_to_excel(
                f'Results/Performance/cv_ci_results_{data_type}_{prediction_type}.xlsx', self.results_df_with_ci)
            self.results_df_with_ci.to_excel(
                f'Results/Performance/cv_ci_results_{data_type}_{prediction_type}_{model_name}.xlsx')

        # possible store the interval table to latex
        if to_latex:
            self.results_df_with_ci.to_latex(
                f'Results/Performance/cv_ci_results_{data_type}_{prediction_type}_{model_name}.tex')

    def generate_hpt_report(self, model_name, data_type, prediction_type):
        """Collect the hyperparametering tuning results and store them to excel

        Args:
            model_name (str): the name of the model for which the hyperparameters were tuned
            data_type (str): knee or hip
            prediction_type (str): diagnosis and prediction
        """

        # initialise the table to hold the results
        hpt_report = pd.DataFrame(columns=['Model', 'Parameter', 'Value'])

        # retrieve the parameter names
        params = self.best_params.keys()
        # retrieve the parameter values
        values = self.best_params.values()

        # retrieve the names of the models for which hyperparameters were tuned
        model_names = [self.model_name for _ in range(
            0, len(self.best_params))]

        # add the parameter names, parameter values, and model names to the table
        hpt_report['Model'] = model_names
        hpt_report['Parameter'] = params
        hpt_report['Value'] = values

        # set multiindex for nice representation of table
        hpt_report.set_index(['Model', 'Parameter'])

        # store the table to excel and latex
        hpt_report.to_excel(
            f'Results/Params/hpt_{data_type}_{prediction_type}_{model_name}.xlsx')
        hpt_report.to_latex(
            f'Results/Params/hpt_{data_type}_{prediction_type}_{model_name}.tex')
        append_df_to_excel(
            f'Results/Params/hpt_{data_type}_{prediction_type}.xlsx', hpt_report)

    def generate_fs_report(self, model_name, data_type, prediction_type):
        """store the feature selection results in a table and store it to excel

        Args:
            model_name (str): the name of the model for which to store the feature selection results
            data_type (str): knee or hip
            prediction_type (str): diagnosis or prediction
        """

        # initialise the table in which the results will be stored
        fs_report = pd.DataFrame(columns=['Threshold', 'Balanced Accuracy'])
        # add the results to the table
        fs_report['Threshold'] = self.fs_thresholds
        fs_report['Balanced Accuracy'] = self.fs_performance
        # store the table to excel
        fs_report.to_excel(
            f'Results/Feature selection/fs_{model_name}_{data_type}_{prediction_type}.xlsx')


@dataclass
class Model:
    """Class to hold a model and its results
    """
    # the machine learning algorithm
    predictor: Any
    # the hyperparameters of the algorithms that can be tuned
    tunable_params: dict
    # general parameters
    params: Any
    # name of the machine learning algorithm
    name: str = None

    def __post_init__(self):
        """Initialise the results of the algorithm
        """
        self.results = Results(self.name)

    def train(self, X_train, y_train):
        """Train the machine learning algorithm

        Args:
            X_train (numpy array): the data to train the model
            y_train (numpy array): the true labels of the training data
        """

        classifier = self.predictor

        # determine which evaluation strategy to use
        if self.params.OVO:
            classifier = OneVsOneClassifier(classifier)
        if self.params.OVR:
            classifier = OneVsRestClassifier(classifier)
        if self.params.CC:
            classifier = CalibratedClassifierCV(
                base_estimator=classifier, cv=self.params.CV_FOLDS)

        # train the algorithm
        classifier.fit(X_train, y_train)
        self.fitted_estimator = classifier

    def save(self):
        """Pickle the model to be used in production
        """
        dump(self.fitted_estimator, open(
            f'Results/Models/{self.name}_model__{self.params.DATA_TYPE}_{self.params.PREDICTION_TYPE}.pkl', 'wb'))

    def predict_test(self, X_test, y_test):
        """Predict the holdout test data

        Args:
            X_test (numpy array): test data
            y_test (numpy array): true labels of the test data

        Returns:
            list, list: label predictions, probability predictions
        """
        self.X_test = X_test
        self.y_test = y_test

        self.y_pred = self.fitted_estimator.predict(self.X_test)

        if self.name != 'ridge':
            self.y_proba = self.fitted_estimator.predict_proba(self.X_test)
        else:
            d = self.fitted_estimator.decision_function(self.X_test)[0]
            self.y_proba = np.exp(d) / np.sum(np.exp(d))

        return self.y_pred, self.y_proba

    def predict_validation(self, data):
        """Predict the validation data

        Args:
            data (class instance): dataset

        Returns:
            list, list: label predictions, probability predictions
        """
        self.X_validation = data.X_validation
        self.y_validation = data.y_validation

        self.y_pred = self.fitted_estimator.predict(data.X_validation)

        if self.name != 'ridge':
            self.y_proba = self.fitted_estimator.predict_proba(
                data.X_validation)
        else:
            d = self.fitted_estimator.decision_function(data.X_validation)[0]
            self.y_proba = np.exp(d) / np.sum(np.exp(d))

        return self.y_pred, self.y_proba

    def set_decision_threshold(self):
        """Optimize the decision threshold of the model and predict test data with new threshold

        Returns:
            list : predictions of the test data with new decision threshold
        """
        # predict the probabilities of the test data
        df = self.fitted_estimator.decision_function(self.X_test)

        # Set the value of decision threshold.
        decision_threshold = self.optimize_threshold()

        # Desired prediction to increase precision value.
        desired_predict = []

        # Iterate through each value of decision function output
        # and if  decision score is > than Decision threshold then,
        # append (1) to the empty list ( desired_prediction) else
        # append (0).
        for i in df:
            if i < decision_threshold:
                desired_predict.append(0)
            else:
                desired_predict.append(1)

        return desired_predict

    def optimize_threshold(self):
        """Optimize the model's decision threshold

        Returns:
            float: the optimal decision threshold
        """
        # predict probabilities
        yhat = self.fitted_estimator.predict_proba(self.X_test)
        # keep probabilities for the positive outcome only
        probs = yhat[:, 1]
        # define thresholds
        thresholds = np.arange(0, 1, 0.001)
        # evaluate each threshold
        scores = [balanced_accuracy_score(
            self.y_test, self.to_labels(probs, t)) for t in thresholds]
        # get best threshold
        ix = np.argmax(scores)

        print(
            f'Threshold : {round(thresholds[ix], 3)}, BA: {round(scores[ix], 3)}')
        self.optimal_theshold = thresholds[ix]
        return thresholds[ix]

    # apply threshold to positive probabilities to create labels
    @staticmethod
    def to_labels(pos_probs, threshold):
        """Apply threshold to positive probabilities to create labels

        Args:
            pos_probs (list): probabilities of instances in the positive class
            threshold (float): the decision threshold

        Returns:
            list: labels
        """
        return (pos_probs >= threshold).astype('int')

    def error_analysis(self, data, columns):
        """Analyse the errors of the model

        Args:
            data (class instance): the dataset
            columns (list): names of the columns
        """

        # convert data to numpy if necessary
        if not self.params.NORMALIZE_DATA and not self.params.STANDARDIZE_DATA:
            data.X_test = data.X_test.to_numpy()
            data.y_test = data.y_test.to_numpy()

        # fit a Model Performance Predictor on the model performances
        error_analyzer = ErrorAnalyzer(
            self.fitted_estimator, feature_names=columns)
        error_analyzer.fit(data.X_test, data.y_test)

        # print metrics regarding the Model Performance Predictor
        print(error_analyzer.mpp_summary(
            data.X_test, data.y_test, output_dict=False))

        # plot the Model Performance Predictor Decision Tree
        error_visualizer = ErrorVisualizer(error_analyzer)
        error_visualizer.plot_error_tree()

        # print the details regarding the decision tree nodes containing the majority of errors
        error_analyzer.error_node_summary(
            leaf_selector="all_errors", add_path_to_leaves=True, print_summary=True)
        # plot the feature distributions of samples in the nodes containing the majority of errors
        # rank features by correlation to error
        error_visualizer.plot_feature_distributions_on_leaves(
            leaf_selector="all_errors", top_k_features=3)

    def explain(self, data, features):
        """Explain the decisions of the model with SHAP analysis

        Args:
            data (class instance): the dataset
            features (list): names of the features in the dataset
        """
        train_data = pd.DataFrame(data, columns=features)
        explainer = shap.KernelExplainer(
            self.fitted_estimator.predict_proba, train_data, link="logit")
        dump(explainer, open(
            f'Results/Transformers/{self.model_name}_{self.params.DATA_TYPE}_{self.params.PREDICTION_TYPE}_explainer.pkl', 'wb'))


@dataclass
class Model_list:
    """Data class to hold a list of models and generate the results of the models in it
    """

    model_list: InitVar
    available_models: InitVar
    tunable_params: InitVar
    params: InitVar

    def __post_init__(self, model_list, available_models, tunable_params, params):
        """Initialise the list of models

        Args:
            model_list (list): the list of models
            available_models (list): a list of the models that are available
            tunable_params (dict): grid with hyperparameters that can be tuned
            params (dict): general training parameters
        """
        # general parameters
        self.params = params
        # initialise the list of model
        self.models = self.initialize_models(
            model_list, available_models, tunable_params, params)

    @staticmethod
    def initialize_models(model_list, available_models, tunable_params, params):
        return [Model(getattr(available_models, model_name), getattr(tunable_params, model_name), params, model_name) for model_name in model_list]

    def generate_hpt_report(self):
        """Save the hyperparameter tuning results for each model in the list of models
        """

        # initialise results table
        hpt_report = pd.DataFrame(columns=['Model', 'Parameter', 'Value'])

        # initialise lists that will hold the results
        params = []
        values = []
        model_names = []

        # for each model in the list of models
        for model in self.models:
            # print the tuning results
            print(model.results.best_params)
            print(model.results.best_params.keys())
            print(model.results.best_params.values())

            # add the parameters, their values, and the model name to the lists that hold the results
            params += list(model.results.best_params.keys())
            values += list(model.results.best_params.values())
            model_names += [model.name for _ in range(
                0, len(model.results.best_params))]

        # add the result lists to the table
        hpt_report['Model'] = model_names
        hpt_report['Parameter'] = params
        hpt_report['Value'] = values

        # set multiindex for nice representation in table
        hpt_report.set_index(['Model', 'Parameter'])

        # store the table to excel and latex
        hpt_report.to_excel(
            f'Results/Params/hpt_{self.params.DATA_TYPE}_{self.params.PREDICTION_TYPE}_{self.params.VALIDATION_OR_TEST}.xlsx')
        hpt_report.to_latex(
            f'Results/Params/hpt_{self.params.DATA_TYPE}_{self.params.PREDICTION_TYPE}_{self.params.VALIDATION_OR_TEST}.tex')

    def confidence_intervals_to_table(self, to_excel=True, to_latex=True):
        """Compute the confidence intervals for each model in the list of models and store them in excel and/or latex

        Args:
            to_excel (bool, optional): Whether to store the results to excel. Defaults to True.
            to_latex (bool, optional): Whether to store the results in latex. Defaults to True.
        """

        # initialise dict that will hold the confidence intervals
        result_dict = {}

        # for each model in the list of models
        for model in self.models:

            # initialise a nested dict that will hold the confidence intervals for this model
            result_dict[model.name] = {}

            # for each metric's confidence interval in the list of the model's metric confidence intervals
            for confidence_interval in model.results.confidence_intervals:

                # retrieve the confidence interval and round it's values
                metric = confidence_interval.metric
                mean = round(confidence_interval.mean, 3)
                left_limit = round(confidence_interval.left_limit, 3)
                right_limit = round(confidence_interval.right_limit, 3)

                # add the confidence interval to the results dict
                result_dict[model.name][metric] = f'{mean} ({left_limit} - {right_limit})'

        # transform the dict to a table
        self.results_df_with_ci = pd.DataFrame.from_dict(
            result_dict, orient='index')
        print(self.results_df_with_ci)

        # possibly store the table to excel
        if to_excel:
            self.results_df_with_ci.to_excel(
                f'Results/Performance/cv_ci_results_{self.params.DATA_TYPE}_{self.params.PREDICTION_TYPE}.xlsx')

        # possible store the table in latex
        if to_latex:
            self.results_df_with_ci.to_latex(
                f'Results/Performance/cv_ci_results_{self.params.DATA_TYPE}_{self.params.PREDICTION_TYPE}.tex')

    def cv_or_tts_results_to_table(self, cv_or_tts='cv', to_excel=True, to_latex=True):
        """Retrieve the metric scores for each model in the list of models and store them to excel and/or latex

        Args:
            cv_or_tts (str, optional): Whether the results were obtained with cross-validation or train_test_split. Defaults to 'cv'.
            to_excel (bool, optional): Whether to store the results in excel. Defaults to True.
            to_latex (bool, optional): Whether to store the data to latex. Defaults to True.
        """

        # obtain the metric scores for each model in the list of models
        results = [model.results.score_dict for model in self.models]

        # create a table with the metric scores of each model
        self.cv_tts_results_df = pd.concat(
            [pd.DataFrame(l) for l in results], axis=1).T
        print(self.cv_tts_results_df)

        # possibly store the table to excel
        if to_excel:
            self.cv_tts_results_df.to_excel(
                f'Results/Performance/{cv_or_tts}_{self.params.VALIDATION_OR_TEST}_results_{self.params.DATA_TYPE}_{self.params.PREDICTION_TYPE}.xlsx')

        # possible store the table in latex
        if to_latex:
            self.cv_tts_results_df.to_latex(
                f'Results/Performance/{cv_or_tts}_{self.params.VALIDATION_OR_TEST}_results_{self.params.DATA_TYPE}_{self.params.PREDICTION_TYPE}.tex')


@dataclass
class Preprocess:
    """A class holding preprocessing parameters with functions to preprocess the data
    """

    # general parameters
    params: Any
    # whether to standardize the data or not
    to_standardize: bool = False
    # whether to normalize the data or not
    to_normalize: bool = False
    # whether to apply PCA to the data or not
    to_pca: bool = False
    # whether to add polynomial features to the data or not
    to_polynomial: bool = False
    # whether to power the data or not
    to_power: bool = False

    @staticmethod
    def standardize(X_train, X_test):
        standardizer = StandardScaler().fit(X_train)
        X_train = standardizer.transform(X_train)
        X_test = standardizer.transform(X_test)
        return X_train, X_test

    @staticmethod
    def normalize(X_train, X_test):
        normalizer = MinMaxScaler().fit(X_train)
        X_train = normalizer.transform(X_train)
        X_test = normalizer.transform(X_test)
        return X_train, X_test

    def PCA(self, X_train, X_test):
        pca_transformer = PCA(self.params.PCA_VARIANCE).fit(X_train)
        X_train = pca_transformer.transform(X_train)
        X_test = pca_transformer.transform(X_test)
        return X_train, X_test

    def polynomial(self, X_train, X_test):
        polynomial_transformer = PolynomialFeatures(
            self.params.AMT_POLY, interaction_only=True).fit(X_train)
        X_train = polynomial_transformer.transform(X_train)
        X_test = polynomial_transformer.transform(X_test)
        return X_train, X_test

    @staticmethod
    def power_transform_data(X_train, X_test):
        power_transformer = power_transform().fit(X_train)
        X_train = power_transformer.transform(X_train)
        X_test = power_transformer.transform(X_test)
        return X_train, X_test


@dataclass
class TrainValidationSplit:
    """Data class to hold and preprocess a split dataset
    """
    # general parameters
    params: Any
    # the training data
    X_train: Any
    # the validation data
    X_validation: Any
    # the training labels
    y_train: Any
    # the validation labels
    y_validation: Any

    available_samplers: InitVar

    def __post_init__(self, available_samplers):
        """Initialise the preprocessing techniques

        Args:
            available_samplers (set): the set of available sampling techniques
        """

        # retrieve preprocessing methods
        self.preprocessing = Preprocess(self.params)

        # possibly set the sampling technique to use
        if self.params.SAMPLING_METHOD is not None:
            strategy_dict = {key: value for key, value in available_samplers.__dict__.items(
            ) if not key.startswith('__') and not callable(key)}
            if self.params.SAMPLING_METHOD == 'both':
                self.oversampler = strategy_dict['over']
                self.undersampler = strategy_dict['under']
                self.sampler = 'both'
            else:
                self.sampler = strategy_dict[self.params.SAMPLING_METHOD]
        else:
            self.sampler = None

    def preprocess(self):
        """Apply the predefined preprocessing techniques to the data
        """
        if self.params.STANDARDIZE_DATA:
            self.X_train, self.X_validation = self.preprocessing.standardize(
                self.X_train, self.X_validation)
        if self.params.NORMALIZE_DATA:
            self.X_train, self.X_validation = self.preprocessing.normalize(
                self.X_train, self.X_validation)
        if self.params.PCA_DATA:
            self.X_train, self.X_validation = self.preprocessing.PCA(
                self.X_train, self.X_validation)
        if self.params.POLY_DATA:
            self.X_train, self.X_validation = self.preprocessing.polynomial(
                self.X_train, self.X_validation)
        if self.params.POWER_DATA:
            self.X_train, self.X_validation = self.preprocessing.power_transform_data(
                self.X_train, self.X_validation)
        if self.sampler is not None:
            try:
                self.resample()
            except AttributeError:
                pass

    def resample(self):
        """Resample the data
        """
        if isinstance(self.sampler, str):
            self.X_train, self.y_train = self.oversampler.fit_resample(
                self.X_train, self.y_train)
            self.X_train, self.y_train = self.undersampler.fit_resample(
                self.X_train, self.y_train)
        else:
            self.X_train, self.y_train = self.sampler.fit_resample(
                self.X_train, self.y_train)


@dataclass
class TunableParameters:
    """Data class that hold the tunable hyperparameters and possible values for each model
    """

    tunable_models: Set[str] = field(default_factory=lambda: {
                                     'ridge', 'svm', 'knn', 'dt', 'rf', 'adaboost', 'nn', 'gb'})

    svm_predictor: Dict[str, list] = field(default_factory=lambda: ({
        'svc__C': ([0.01, 0.1, 1]),
        'svc__kernel': (['rbf', 'linear', 'sigmoid']),
        'svc__degree': ([1, 3]),
        'svc__decision_function_shape': (['ovo', 'ovr']),
    }))

    dt: Dict[str, list] = field(default_factory=lambda: ({
        'decisiontreeclassifier__criterion': (['gini', 'entropy']),
        'decisiontreeclassifier__max_depth': ([10, 20]),
        'decisiontreeclassifier__min_samples_split': ([2, 3, 5]),
        'decisiontreeclassifier__min_samples_leaf': ([2, 3, 5]),
    }))

    rf: Dict[str, list] = field(default_factory=lambda: ({
        # 'randomforestclassifier__bootstrap': [True, False],
        'randomforestclassifier__n_estimators': ([50, 100, 200, 400]),
        # 'randomforestclassifier__max_features': (['auto', 'sqrt']),
        'randomforestclassifier__max_depth':    ([1, 3, 5, 10, 20, None]),
        'randomforestclassifier__min_samples_split':  [2, 5, 10],
        'randomforestclassifier__min_samples_leaf':   [1, 2, 4, 10],
    }))

    ridge: Dict[str, list] = field(default_factory=lambda: ({
        'alpha': ([0.1, 0.2, 0.4, 0.8, 1, 2, 4, 8, 16, 32, 64, 128]),
        'solver': (['auto', 'svd', 'cholesky', 'lsqr', 'sparse_cg', 'sag', 'saga'])
    }))

    knn: Dict[str, list] = field(default_factory=lambda: ({
        'kneighborsclassifier__weights': (['uniform', 'distance']),
        'kneighborsclassifier__n_neighbors': ([2, 3, 5, 10, 15, 20]),
        'kneighborsclassifier__algorithm': (['ball_tree', 'kd_tree', 'brute']),
        'kneighborsclassifier__leaf_size': ([10, 20, 30, 60, 90]),
        'kneighborsclassifier__p': ([1, 2])
    }))

    adaboost: Dict[str, list] = field(default_factory=lambda: ({
        'adaboostclassifier__n_estimators': ([20, 40, 50, 100, 200, 400, 800]),
        'adaboostclassifier__learning_rate': ([0.001, 0.01, 0.1]),
    }))

    gb: Dict[str, list] = field(default_factory=lambda: ({
        'n_estimators': ([50, 100, 200, 400, 800]),
        'learning_rate': ([0.001, 0.01, 0.1, 1]),
        'min_samples_split':  [2, 5, 10],
        'min_samples_leaf':   [1, 2, 4, 10],
        'max_depth':    ([1, 3, 5, 10, 20, 40, 80]),
    }))

    nn: Dict[str, list] = field(default_factory=lambda: ({
        'activation': (['identity', 'logistic', 'tanh', 'relu']),
        'solver': (['lbfgs', 'adam']),
        'alpha': ([0.0001, 0.001, 0.01, 0.1]),
        'learning_rate': (['constant', 'invscaling', 'adaptive']),
        'early_stopping': ([True]),
        'max_iter': ([500, 1000])
    }))

    sc: Dict[str, list] = field(default_factory=lambda: ({
    }))

    def check_validity(self, model_list):
        if any(model not in self.tunable_models for model in model_list):
            raise ValueError(
                f'Chosen model(s) must be one or more of {self.tunable_models}')


@dataclass
class ModelOptions:
    """Data class that holds the models that can be chosen for training and evaluation
    """

    svm_predictor: Any = svm.SVC(class_weight='balanced', probability=True,
                                 kernel='linear', degree=1, decision_function_shape='ovo', C=0.1)
    knn: Any = neighbors.KNeighborsClassifier()
    dt: Any = DecisionTreeClassifier(class_weight='balanced')
    rf: Any = RandomForestClassifier(class_weight='balanced', n_estimators=100, min_samples_split=2, min_samples_leaf=1, max_features='auto',
                                     max_depth=3, bootstrap=False)
    adaboost: Any = AdaBoostClassifier(n_estimators=800, learning_rate=0.13)
    gb: Any = GradientBoostingClassifier(
        n_estimators=2000, min_samples_split=5, min_samples_leaf=1, max_depth=3, learning_rate=0.1)
    nn: Any = MLPClassifier(solver='lbfgs', alpha=1e-5,
                            hidden_layer_sizes=(5, 2))
    sc: Any = StackingClassifier(estimators=[('svm_predictor', svm.SVC(class_weight='balanced')),
                                             ('ridge', RidgeClassifier(
                                                 class_weight='balanced')),
                                             ('knn', neighbors.KNeighborsClassifier()),
                                             ('dt', DecisionTreeClassifier(
                                                 class_weight='balanced')),
                                             ('rf', RandomForestClassifier(
                                                 class_weight='balanced')),
                                             ('adaboost', AdaBoostClassifier()),
                                             ('gb', GradientBoostingClassifier())],
                                 final_estimator=LogisticRegression(max_iter=1000, class_weight='balanced', solver='saga', n_jobs=-1))

    available: Set[str] = field(default_factory=lambda: {
                                'ridge', 'svm_predictor', 'knn', 'dt', 'rf', 'adaboost', 'nn', 'gb', 'sc'})

    def check_validity(self, model_list):
        print(model_list)
        if any(model not in self.available for model in model_list):
            raise ValueError(
                f'Chosen model(s) must be one or more of {self.available}')


@dataclass
class SamplingMethods:
    """Data class that holds the available sampling methods

    Raises:
        ValueError: if a method is chosen that is not available
    """
    params: Any

    # list with the available sampling methods
    available: Set[str] = field(default_factory=lambda: {
                                'under', 'over', 'smote', 'smotetomek', 'both', None, 'nearmiss'})

    def __post_init__(self):
        """Initialise the available sampling methods
        """
        # smote oversampling
        self.smote = SMOTE()
        # random undersampling
        self.under = RandomUnderSampler(
            sampling_strategy=self.params.UNDERSAMPLE_STRATEGY, random_state=42)
        # random oversampling
        self.over = RandomOverSampler(
            sampling_strategy=self.params.OVERSAMPLE_STRATEGY, random_state=42)
        # combination of smote and tomek links
        self.smotetomek = SMOTETomek(sampling_strategy='all')
        # nearmiss
        self.nearmiss = NearMiss(version=1, sampling_strategy='all')

    def check_validity(self, method):
        if method not in self.available:
            raise ValueError(
                f'Sampling method must be one of {self.available}')


@dataclass
class FeatureImportances:
    """Data class that with feature selection methods, their results, and functions to plot the results
    """
    params: Any
    # lists that will hold the feature importances computed by each feature selection technique
    dt: list = field(default_factory=list)
    rf: list = field(default_factory=list)
    ridge: list = field(default_factory=list)
    permutation: list = field(default_factory=list)
    kbest: list = field(default_factory=list)
    rfe: list = field(default_factory=list)

    def compute_ridge_feature_importances(self, X, y, columns, model_name):
        """Compute ridge feature importance

        Args:
            X (numpy array): the data
            y (numpy array): true labels of the data
            columns (list): column names
            model_name (str): name of the model for which importances are computed
        """

        print('Computing ridge feature importances.........')

        # repeat for N_rounds
        for i in range(self.params.N_ROUNDS):
            # initialise ridge classifier
            ridge_model = RidgeClassifier(class_weight='balanced')
            # fit the model to the data
            ridge_model.fit(X, y)
            # retrieve the coëfficients of the model
            feature_importances = ridge_model.coef_
            # add them to the list of feature importances
            self.ridge.append(list(np.mean(feature_importances, axis=0)))

        # average the importances from all N_rounds
        averaged_feature_importances = np.mean(self.ridge, axis=0)
        self.ridge = np.abs(averaged_feature_importances)
        # plot the averaged feature importances
        self.plot_feature_importances(
            columns, averaged_feature_importances, 'ridge', model_name)

    def compute_rf_feature_importance(self, X, y, columns, model_name):
        """Compute ridge feature importance

        Args:
            X (numpy array): the data
            y (numpy array): true labels of the data
            columns (list): column names
            model_name (str): name of the model for which importances are computed
        """

        print('Computing random forest feature importances.........')

        # repeat for N_ROUNDS
        for i in range(self.params.N_ROUNDS):
            # initialise the random forest classifier
            rf_model = RandomForestClassifier(class_weight='balanced')
            # fit the model to the data
            rf_model.fit(X, y)
            # retrieve it's feature importance scores
            feature_importances = rf_model.feature_importances_
            # add the scores to the list of scores
            self.rf.append(feature_importances)

        # average the scores from all N_ROUNDS
        averaged_feature_importances = np.mean(self.rf, axis=0)
        self.rf = averaged_feature_importances

        # plot the feature importance scores
        self.plot_feature_importances(
            columns, averaged_feature_importances, 'rf', model_name)

    def compute_permutation_feature_importance(self, X, y, columns, model_name):
        """Compute ridge feature importance

        Args:
            X (numpy array): the data
            y (numpy array): true labels of the data
            columns (list): column names
            model_name (str): name of the model for which importances are computed
        """

        print('Computing permutation feature importances.........')

        # repeat for N_ROUNDS
        for i in range(self.params.N_ROUNDS):
            # initialise the random forest classifier
            rf_model = RandomForestClassifier(class_weight='balanced')
            # fit the model the data
            rf_model.fit(X, y)
            # perform permutation feature importance
            results = permutation_importance(
                rf_model, X, y, scoring='balanced_accuracy')
            # retrieve the feature importance scores
            feature_importances = results.importances_mean
            # add them to the lest of feature importance scores
            self.permutation.append(feature_importances)

        # average the feature importance scores for all N_ROUNDS
        averaged_feature_importances = np.mean(self.permutation, axis=0)
        self.permutation = averaged_feature_importances

        # plot the feature importance scores
        self.plot_feature_importances(
            columns, averaged_feature_importances, 'permutation', model_name)

    def compute_Kbest_feature_importance(self, X, y, columns, model_name):
        """Compute ridge feature importance

        Args:
            X (numpy array): the data
            y (numpy array): true labels of the data
            columns (list): column names
            model_name (str): name of the model for which importances are computed
        """

        print('Computing Kbest feature importances........')

        # repeat for N_ROUNDS
        for i in range(self.params.N_ROUNDS):
            # initialise Kbest feature selection
            bestfeatures = SelectKBest(
                score_func=mutual_info_classif, k=len(columns))
            # fit Kbest to the data
            fit = bestfeatures.fit(X, y)
            # retrieve the feature importance scores
            feature_importances = fit.scores_
            # add the feature importance scores to the list of scores
            self.kbest.append(feature_importances)

        # average the feature importance scores for all N_ROUNDS
        averaged_feature_importances = np.mean(self.kbest, axis=0)
        self.kbest = averaged_feature_importances

        # plot the averaged feature importance scores
        self.plot_feature_importances(
            columns, averaged_feature_importances, 'kbest', model_name)

    def combine_fi_methods(self, columns, model_name):
        """Combine multiple feature selection methods to select the best subset of features

        Args:
            columns (list): column names
            model_name (str): name of the model for which features are selected
        """

        # compute random forest feature importance scores
        rf_importances = self.compute_normalized(self.rf)
        # compute permutation feature importance scores
        permutation_importances = self.compute_normalized(self.permutation)
        # compute Kbest feature importance scores
        kbest_importances = self.compute_normalized(self.kbest)
        # perform ridge feature importance scores
        ridge_importances = self.compute_normalized(self.ridge)
        # compute the feature importance scores by combining the scores of previous methods
        combined_importances = [self.compute_magnitude(scores) for scores in zip(
            rf_importances, permutation_importances, kbest_importances, ridge_importances)]  # , rfe_importances)]
        # plot the feature importance scores
        self.plot_feature_importances(
            columns, combined_importances, 'combined', model_name)

    def compute_normalized(self, importances, inverted=False):
        """Normalize feature importance scores to 0-1 range

        Args:
            importances (list): feature importance scores
            inverted (bool, optional): Whether lower scores are better or not. Defaults to False.

        Returns:
            list: normalized feature importance scores
        """
        max_importance = max(importances)
        normalized_importances = [self.normalize_fi(
            score, max_importance, inverted) for score in importances]
        return normalized_importances

    def normalize_fi(self, score, max_fi, inverted):
        normalized_score = (score / max_fi) ** 2
        if inverted:
            normalized_score = 1 - normalized_score
        return normalized_score

    def compute_magnitude(self, scores):
        return np.sqrt(sum(list(scores)))

    def plot_feature_importances(self, columns, importances, method, model_name, figsize=(10, 10), title="Feature Importances"):
        """Plot the feature importances of a given feature importance
        computation method.

        Args:
            importances (array): feature importancesch importances are computed. Defaults to 'all_classes'.
            figsize (tuple, optional): the size of the plot. Defaults to (10,10).
            title (str, optional): The title of the plot. Defaults to "Feature Importances".
            top_n (int, optional): The amount of top features to plot. Defaults to 10.
        """
        feat_imp = pd.DataFrame({'importance': importances})
        feat_imp['feature'] = columns
        feat_imp.sort_values(by='importance', ascending=False, inplace=True)
        self.combined_importances = feat_imp
        feat_imp.to_excel(
            f'Data/feature_importances_{method}_{self.params.DATA_TYPE}_{self.params.PREDICTION_TYPE}.xlsx')
        feat_imp = feat_imp.iloc[:self.params.FI_TOP_N]

        feat_imp.sort_values(by='importance', inplace=True)
        feat_imp = feat_imp.set_index('feature', drop=True)
        feat_imp.plot.barh(title=title, figsize=figsize)
        plt.xlabel('Feature Importance Score')

        plt.savefig(
            f'Results/Plots/feature_importances/{method}_fi_{self.params.DATA_TYPE}_{self.params.PREDICTION_TYPE}_{model_name}.png')
        plt.close()


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """

    import os

    from openpyxl import load_workbook

    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
